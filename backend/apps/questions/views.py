import io
from django.db import transaction
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from common.response import api_response
from common.exceptions import BusinessError
from apps.courses.models import Course
from apps.knowledge.models import Chapter, KnowledgePoint
from .models import GenerationTask, Question
from .serializers import GenerationTaskSerializer, QuestionSerializer, QuestionReviewSerializer
from .services import ai_review, find_similar, question_hash

class GenerationTaskViewSet(viewsets.ModelViewSet):
    serializer_class = GenerationTaskSerializer
    queryset = GenerationTask.objects.select_related("course").order_by("-created_at", "-id")
    http_method_names = ["get", "post", "delete"]
    def create(self, request, *args, **kwargs):
        course = get_object_or_404(Course, pk=request.data.get("course"), is_deleted=False)
        count = int(request.data.get("count", 1))
        if count < 1 or count > 100: raise BusinessError("一次任务的题目数量应为1到100。", 40043)
        config = request.data.copy()
        question_types = config.get("question_types") or ["single_choice"]
        if not isinstance(question_types, list) or not question_types:
            raise BusinessError("请至少选择一种题型。", 40047)
        from common.constants import QUESTION_TYPES
        unsupported = [item for item in question_types if item not in QUESTION_TYPES]
        if unsupported: raise BusinessError(f"不支持的题型：{', '.join(unsupported)}", 40048)
        type_counts = config.get("type_counts") or {}
        if type_counts:
            if not isinstance(type_counts, dict): raise BusinessError("题型数量格式不正确。", 40049)
            try:
                normalized_counts = {key: int(value) for key, value in type_counts.items() if int(value) > 0}
            except (TypeError, ValueError):
                raise BusinessError("每种题型的数量必须是正整数。", 40050)
            invalid_types = [key for key in normalized_counts if key not in QUESTION_TYPES]
            if invalid_types: raise BusinessError(f"不支持的题型：{', '.join(invalid_types)}", 40048)
            if sum(normalized_counts.values()) != count: raise BusinessError("各题型数量之和必须等于生成总数。", 40051)
            config["type_counts"] = normalized_counts
            config["question_types"] = list(normalized_counts)
        config["generate_answer"] = True
        config["generate_analysis"] = True
        quality_mode = str(config.get("quality_mode", "STANDARD")).upper()
        if quality_mode not in {"FAST", "STANDARD", "DEEP"}:
            raise BusinessError("质量模式只能是快速、标准或深度模式。", 40052)
        config["quality_mode"] = quality_mode
        task = GenerationTask.objects.create(course=course, config=dict(config), total_count=count, status="WAITING")
        # 只有深度模式才启用完整的多智能体协作；快速和标准模式直接使用生成服务。
        if quality_mode == "DEEP":
            from apps.agents.services import create_question_workflow
            create_question_workflow(task, quality_mode)
        return api_response(GenerationTaskSerializer(task).data, "出题任务已创建，后台Worker将自动处理。", request=request)

    def destroy(self, request, *args, **kwargs):
        task = self.get_object()
        if task.status in {"WAITING", "RUNNING"}:
            raise BusinessError("任务仍在执行，请先取消任务，等待状态变为已取消后再删除。", 40922, 409)
        from apps.agents.models import AgentWorkflowRun
        AgentWorkflowRun.objects.filter(business_type="generation_task", business_id=task.id).delete()
        # Question.generation_task 使用 SET_NULL；这里只删除任务记录，已生成题目继续保留在题库。
        task.delete()
        return api_response(None, "生成任务记录已删除，已生成题目继续保留。", request=request)
    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        from apps.agents.models import AgentWorkflowRun
        task = self.get_object()
        if task.status not in {"SUCCESS", "FAILED", "CANCELLED", "INTERRUPTED"}:
            finished_at = timezone.now()
            # 页面立即显示已取消；正在进行的模型请求返回后，也会通过
            # cancel_requested 停止后续步骤，不再继续生成或审核。
            GenerationTask.objects.filter(pk=task.pk).update(
                cancel_requested=True,
                status="CANCELLED",
                finished_at=finished_at,
            )
            AgentWorkflowRun.objects.filter(
                business_type="generation_task",
                business_id=task.id,
                status__in=["WAITING", "RUNNING"],
            ).update(
                cancel_requested=True,
                status="CANCELLED",
                finished_at=finished_at,
            )
            task.refresh_from_db()
        return api_response(GenerationTaskSerializer(task).data, "任务已取消", request=request)
    @action(detail=True, methods=["post"])
    def retry(self, request, pk=None):
        task = self.get_object()
        if task.status not in ["FAILED", "CANCELLED", "INTERRUPTED"]: raise BusinessError("只有失败、取消或中断的任务可以重试。", 40921, 409)
        actual_count = min(task.total_count, task.questions.filter(is_deleted=False).count())
        task.status = "WAITING"; task.progress = int(actual_count / task.total_count * 100) if task.total_count else 0
        task.cancel_requested = False; task.error_message = ""; task.success_count = actual_count
        task.failed_count = max(0, task.total_count - actual_count)
        task.config = {key: value for key, value in task.config.items() if key != "failure_details"}
        task.save()
        from apps.agents.models import AgentWorkflowRun
        workflow = AgentWorkflowRun.objects.filter(business_type="generation_task", business_id=task.id).order_by("-created_at").first()
        if task.config.get("quality_mode") == "DEEP":
            from apps.agents.services import create_question_workflow, retry_workflow
            if workflow and workflow.status in ["FAILED", "CANCELLED", "INTERRUPTED"]: retry_workflow(workflow)
            elif not workflow: create_question_workflow(task, "DEEP")
        elif workflow and workflow.status in ["WAITING", "RUNNING", "FAILED", "INTERRUPTED"]:
            workflow.status = "CANCELLED"; workflow.finished_at = timezone.now(); workflow.save()
        return api_response(GenerationTaskSerializer(task).data, f"任务已重新加入队列，将从{actual_count}道继续补齐", request=request)

    @action(detail=True, methods=["post"], url_path="batch-review")
    def batch_review(self, request, pk=None):
        task = self.get_object()
        mapping = {"approve_ids": "APPROVED", "revision_ids": "NEEDS_REVISION", "reject_ids": "REJECTED"}
        affected = 0
        with transaction.atomic():
            for key, status_value in mapping.items():
                ids = request.data.get(key, []) or []
                affected += task.questions.filter(id__in=ids, is_deleted=False).update(review_status=status_value)
        from apps.agents.models import AgentWorkflowRun
        workflow = AgentWorkflowRun.objects.filter(business_type="generation_task", business_id=task.id).order_by("-created_at").first()
        remaining = task.questions.filter(is_deleted=False, review_status="PENDING").count()
        if workflow:
            workflow.result = {**workflow.result, "human_feedback": str(request.data.get("feedback", ""))[:1000], "remaining_pending": remaining}
            if remaining == 0: workflow.status = "SUCCESS"
            workflow.save()
        return api_response({"affected": affected, "remaining_pending": remaining}, "批量审核已完成", request=request)

class QuestionViewSet(viewsets.ModelViewSet):
    serializer_class = QuestionSerializer
    def get_queryset(self):
        qs = Question.objects.filter(is_deleted=False).select_related("course", "chapter", "knowledge_point").prefetch_related("options", "reviews")
        mapping = {"course": "course_id", "chapter": "chapter_id", "knowledge_point": "knowledge_point_id", "question_type": "question_type", "difficulty": "difficulty", "review_status": "review_status", "source_type": "source_type"}
        for param, field in mapping.items():
            if self.request.query_params.get(param): qs = qs.filter(**{field: self.request.query_params[param]})
        if self.request.query_params.get("keyword"): qs = qs.filter(stem__icontains=self.request.query_params["keyword"])
        return qs
    def perform_destroy(self, instance): instance.is_deleted = True; instance.save()
    def _status(self, request, status_value, message):
        item = self.get_object(); item.review_status = status_value; item.save(); return api_response(QuestionSerializer(item).data, message, request=request)
    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None): return self._status(request, "APPROVED", "题目已审核通过")
    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None): return self._status(request, "REJECTED", "题目已审核不通过")
    @action(detail=True, methods=["post"])
    def revision(self, request, pk=None): return self._status(request, "NEEDS_REVISION", "题目已退回修改")
    @action(detail=True, methods=["post"], url_path="ai-review")
    def ai_review_action(self, request, pk=None): return api_response(QuestionReviewSerializer(ai_review(self.get_object())).data, "AI审核完成", request=request)
    @action(detail=True, methods=["post"])
    def favorite(self, request, pk=None):
        item = self.get_object(); item.is_favorite = not item.is_favorite; item.save(); return api_response({"is_favorite": item.is_favorite}, "收藏状态已更新", request=request)
    @action(detail=False, methods=["post"], url_path="similarity-check")
    def similarity_check(self, request):
        return api_response(find_similar(request.data.get("stem", ""), request.data.get("course_id"), float(request.data.get("threshold", 0.88)), request.data.get("exclude_id")), request=request)
    @action(detail=False, methods=["post"], url_path="batch-action")
    def batch_action(self, request):
        qs = Question.objects.filter(id__in=request.data.get("ids", []), is_deleted=False); action_name = request.data.get("action")
        if action_name == "approve": qs.update(review_status="APPROVED")
        elif action_name == "delete": qs.update(is_deleted=True)
        elif action_name == "difficulty": qs.update(difficulty=request.data.get("value"))
        elif action_name == "score": qs.update(score=request.data.get("value"))
        elif action_name == "chapter": qs.update(chapter_id=request.data.get("value"))
        elif action_name == "knowledge_point": qs.update(knowledge_point_id=request.data.get("value"))
        else: raise BusinessError("不支持的批量操作。", 40044)
        return api_response({"affected": qs.count()}, "批量操作完成", request=request)
    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser], url_path="import")
    def import_questions(self, request):
        from openpyxl import load_workbook
        file = request.FILES.get("file")
        if not file: raise BusinessError("请选择Excel文件。", 40045)
        wb = load_workbook(file, read_only=True, data_only=True); ws = wb.active; headers = [c.value for c in next(ws.iter_rows())]
        required = ["题型", "题干", "正确答案", "难度", "分值", "课程"]
        if any(x not in headers for x in required): raise BusinessError("Excel缺少必要列，请先下载导入模板。", 40046)
        success, failed, duplicate, errors = 0, 0, 0, []
        type_map = {"单项选择题":"single_choice", "多项选择题":"multiple_choice", "判断题":"judge", "填空题":"fill_blank", "简答题":"short_answer", "论述题":"essay", "计算题":"calculation", "编程题":"programming", "案例分析题":"case_analysis", "名词解释题":"term_explanation"}
        for row_no, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            row = dict(zip(headers, cells)); stem = str(row.get("题干") or "").strip()
            try:
                course = Course.objects.get(name=row["课程"], is_deleted=False)
                if Question.objects.filter(course=course, content_hash=question_hash(stem), is_deleted=False).exists(): duplicate += 1; continue
                qtype = type_map.get(row["题型"], row["题型"]); answers = [x.strip().upper() for x in str(row["正确答案"] or "").replace("，", ",").split(",") if x.strip()]
                options = [{"label": label, "content": row.get(f"选项{label}")} for label in "ABCDEF" if row.get(f"选项{label}")]
                serializer = QuestionSerializer(data={"course": course.id, "question_type": qtype, "stem": stem, "options": options, "answer": answers, "analysis": row.get("解析") or "", "scoring_points": [x for x in str(row.get("评分要点") or "").split("|") if x], "difficulty": row["难度"], "score": row["分值"], "review_status": "APPROVED"})
                serializer.is_valid(raise_exception=True); serializer.save(); success += 1
            except Exception as exc: failed += 1; errors.append({"row": row_no, "reason": str(exc)})
        return api_response({"total": success + failed + duplicate, "success": success, "failed": failed, "duplicate": duplicate, "errors": errors}, "导入完成", request=request)
    @action(detail=False, methods=["get"])
    def export(self, request):
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "题库"
        headers = ["题型", "题干", "选项A", "选项B", "选项C", "选项D", "选项E", "选项F", "正确答案", "参考答案", "解析", "评分要点", "难度", "分值", "课程", "章节", "知识点"]
        ws.append(headers)
        if request.query_params.get("template") != "1":
            for q in self.get_queryset()[:5000]:
                option_map = {o.label: o.content for o in q.options.all()}; ws.append([q.question_type, q.stem, *[option_map.get(x, "") for x in "ABCDEF"], ",".join(map(str, q.answer)), ",".join(map(str, q.answer)), q.analysis, "|".join(q.scoring_points), q.difficulty, float(q.score), q.course.name, q.chapter.name if q.chapter else "", q.knowledge_point.name if q.knowledge_point else ""])
        else:
            ws.append(["单项选择题", "示例题干（请删除本行后填写）", "选项A", "选项B", "选项C", "选项D", "", "", "A", "A", "示例解析", "", "中等", 2, "系统中已有课程名称", "", ""])
        stream = io.BytesIO(); wb.save(stream); stream.seek(0)
        filename = "question_import_template.xlsx" if request.query_params.get("template") == "1" else "questions.xlsx"
        return FileResponse(stream, as_attachment=True, filename=filename)
